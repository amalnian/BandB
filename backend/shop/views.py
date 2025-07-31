import datetime
import logging
from django.db import models
from shop import serializers
import jwt
from dateutil.parser import parse

from django.conf import settings
from django.core.mail import send_mail
from django.contrib.auth import authenticate
from django.db.models import Sum
from django.forms import ValidationError
from django.shortcuts import get_object_or_404
from django.utils.translation import gettext_lazy as _
from django.db import transaction
from django.core.cache import cache
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
import logging

from shop.tasks import (
    send_shop_registration_otp_task,
    send_shop_forgot_password_otp_task,
    send_shop_resend_otp_task
)

logger = logging.getLogger(__name__)

from rest_framework import status, permissions, generics
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer, TokenRefreshSerializer
from rest_framework_simplejwt.token_blacklist.models import BlacklistedToken
from rest_framework_simplejwt.tokens import RefreshToken, AccessToken
from rest_framework_simplejwt.views import (
    TokenObtainPairView,
    TokenRefreshView,
)

from chat.models import Notification
from users.models import CustomUser
from shop.models import (
    Booking, BookingFeedback, Shop, ShopCommissionPayment, ShopImage, Service, OTP,
    BusinessHours, SpecialClosingDay
)
from shop.serializers import (
    BookingFeedbackSerializer,
    CustomTokenObtainPairSerializer,
    ShopForgotPasswordSerializer,
    ShopResetPasswordSerializer,
    ShopSerializer,
    ShopImageSerializer,
    # NotificationSerializer,
    ServiceSerializer,
    ShopVerifyForgotPasswordOTPSerializer,
    BusinessHoursSerializer,
    SpecialClosingDaySerializer,
)



logger = logging.getLogger(__name__)



class ShopRegisterView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        # Log the received data for debugging
        logger.debug(f"Received data: {request.data}")
        
        # If the data is structured differently than expected, restructure it
        data = request.data
        if 'user' in data and isinstance(data['user'], dict) and 'shop' in data and isinstance(data['shop'], dict):
            # The frontend is sending nested data, extract and merge it
            user_data = data['user']
            shop_data = data['shop']
            
            # Merge data for the serializer
            serializer_data = {
                'name': shop_data.get('name', ''),
                'email': user_data.get('email', ''),
                'password': user_data.get('password', ''),
                'username': user_data.get('username', ''),
                'phone': shop_data.get('phone', ''),
                'address': shop_data.get('address', ''),
                'description': shop_data.get('description', ''),
                'owner_name': shop_data.get('owner_name', ''),
                'latitude': shop_data.get('latitude'),
                'longitude': shop_data.get('longitude')
            }
        else:
            # Direct data format
            serializer_data = data
        
        # DEBUG: Log coordinate values and their lengths
        lat = serializer_data.get('latitude')
        lng = serializer_data.get('longitude')
        if lat is not None:
            logger.debug(f"Latitude value: {lat}, type: {type(lat)}, length: {len(str(lat))}")
        if lng is not None:
            logger.debug(f"Longitude value: {lng}, type: {type(lng)}, length: {len(str(lng))}")
        
        # Process coordinates to ensure they fit the model constraints
        if lat is not None:
            try:
                # Convert to float first, then to string with limited precision
                lat_float = float(lat)
                # Round to 9 decimal places to fit within 12 total digits
                serializer_data['latitude'] = round(lat_float, 9)
                logger.debug(f"Processed latitude: {serializer_data['latitude']}")
            except (ValueError, TypeError):
                logger.error(f"Invalid latitude value: {lat}")
                return Response(
                    {'detail': 'Invalid latitude value provided.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        if lng is not None:
            try:
                # Convert to float first, then to string with limited precision
                lng_float = float(lng)
                # Round to 9 decimal places to fit within 12 total digits
                serializer_data['longitude'] = round(lng_float, 9)
                logger.debug(f"Processed longitude: {serializer_data['longitude']}")
            except (ValueError, TypeError):
                logger.error(f"Invalid longitude value: {lng}")
                return Response(
                    {'detail': 'Invalid longitude value provided.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Check if email already exists before passing to serializer
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        email = serializer_data.get('email')
        if email and User.objects.filter(email=email).exists():
            return Response(
                {'detail': 'A user with this email already exists.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Now use the processed data
        serializer = ShopSerializer(data=serializer_data)
        
        if serializer.is_valid():
            try:
                # Save the shop with is_active=False
                shop = serializer.save()
                
                # Generate OTP using the OTP model
                otp_obj = OTP.create_for_shop(shop)
                
                send_shop_registration_otp_task.delay(
                    email=shop.email,
                    otp=otp_obj.otp_code,
                    shop_name=shop.name,
                    owner_name=shop.owner_name
                )
                
                logger.info(f"Shop registration OTP email task queued for {shop.email}")
                
                return Response({
                    'success': True,
                    'detail': 'Registration successful. Please check your email for verification code.',
                    'email': shop.email
                }, status=status.HTTP_201_CREATED)
                
            except Exception as e:
                logger.error(f"Error during shop registration: {str(e)}")
                return Response(
                    {'detail': f'Registration failed: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            logger.warning(f"Validation errors: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ShopVerifyOTPView(APIView):
    permission_classes = [AllowAny]
    
    def post(self, request):
        email = request.data.get('email')
        otp_code = request.data.get('otp')
        
        # Add detailed logging
        logger.info(f"OTP verification attempt for email: {email} with code: {otp_code}")
        
        if not email or not otp_code:
            return Response(
                {'detail': 'Email and OTP are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            shop = Shop.objects.get(email=email)
            logger.info(f"Found shop with ID: {shop.id}, name: {shop.name}")
        except Shop.DoesNotExist:
            logger.warning(f"No shop found with email: {email}")
            return Response(
                {'detail': 'Shop not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Enhanced debugging for OTP investigation
        from django.utils import timezone
        current_time = timezone.now()
        logger.info(f"Current time: {current_time}")
        
        # Check all OTPs for this shop (including expired ones)
        all_otps = OTP.objects.filter(shop=shop).order_by('-created_at')
        logger.info(f"Total OTPs ever created for shop {shop.id}: {all_otps.count()}")
        
        for i, otp in enumerate(all_otps[:5]):  # Log first 5 OTPs
            is_expired = otp.expires_at < current_time
            logger.info(f"OTP {i+1}: code={otp.otp_code}, created={otp.created_at}, expires={otp.expires_at}, expired={is_expired}")
        
        # Check current active OTPs using the relationship
        active_otps = shop.otps.all()
        otp_count = active_otps.count()
        logger.info(f"Active OTPs using shop.otps relationship: {otp_count}")
        
        if otp_count == 0:
            # Check if there were ever any OTPs created
            total_otps_ever = OTP.objects.filter(shop=shop).count()
            logger.warning(f"No active OTPs found. Total OTPs ever created: {total_otps_ever}")
            
            if total_otps_ever > 0:
                logger.warning("OTPs were created but are no longer active - possibly deleted by cleanup process")
            
            return Response(
                {'non_field_errors': ['No OTP found for this account. Please request a new one.']},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get the latest active OTP
        try:
            otp_obj = active_otps.latest('created_at')
            logger.info(f"Latest active OTP: code={otp_obj.otp_code}, created={otp_obj.created_at}, expires={otp_obj.expires_at}")
            logger.info(f"OTP is_valid() result: {otp_obj.is_valid()}")
        except OTP.DoesNotExist:
            return Response(
                {'non_field_errors': ['No OTP found for this account. Please request a new one.']},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if OTP is valid and matches
        if not otp_obj.is_valid():
            logger.warning(f"OTP expired. Current time: {current_time}, OTP expires: {otp_obj.expires_at}")
            return Response(
                {'non_field_errors': ['OTP has expired. Please request a new OTP.']},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Convert input OTP to string and strip whitespace for comparison
        if not isinstance(otp_code, str):
            otp_code = str(otp_code)
        otp_code = otp_code.strip()
        stored_otp = otp_obj.otp_code.strip()
        
        logger.info(f"Comparing OTPs - input: '{otp_code}', stored: '{stored_otp}'")
        
        if otp_code != stored_otp:
            logger.warning(f"OTP mismatch: received '{otp_code}', stored '{stored_otp}'")
            return Response(
                {'non_field_errors': ['Invalid OTP. Please try again.']},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # OTP is valid and matches
        logger.info(f"OTP verification successful for shop: {shop.id}")
        
        # Activate the user account
        user = shop.user
        user.is_active = True
        user.save()
        
        # Mark email as verified
        shop.is_email_verified = True
        shop.save()
        
        # Delete the used OTP
        otp_obj.delete()
        logger.info(f"OTP deleted successfully for shop: {shop.id}")
        
        return Response(
            {'detail': 'Email verification successful.'},
            status=status.HTTP_200_OK
        )

class ShopResendOTPView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')
        
        if not email:
            return Response({
                'detail': 'Email is required.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            shop = Shop.objects.get(email=email)
        except Shop.DoesNotExist:
            return Response({
                'detail': 'No shop found with this email.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Generate new OTP
        otp_obj = OTP.create_for_shop(shop)
        
        send_shop_resend_otp_task.delay(
            email=shop.email,
            otp=otp_obj.otp_code,
            shop_name=shop.name,
            owner_name=shop.owner_name
        )
        
        logger.info(f"Shop resend OTP email task queued for {shop.email}")
        
        return Response({
            'success': True,
            'detail': 'New verification code sent to your email.'
        }, status=status.HTTP_200_OK)
    


class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token['role'] = user.role  
        return token


# Updated backend views with fixes

# Updated views.py with fixes

class CustomShopTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
    permission_classes = [permissions.AllowAny]
    
    def post(self, request, *args, **kwargs):
        try:
            email = request.data.get("email")
            password = request.data.get("password")

            # Check if user exists
            if not CustomUser.objects.filter(email=email).exists():
                return Response(
                    {"success": False, "message": "User with this account does not exist"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get the user and check if they have a shop
            user = CustomUser.objects.get(email=email)
            
            # Check if user is active
            if not user.is_active:
                return Response(
                    {"success": False, "message": "Account is not activated. Please verify your email first."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if user is a shop owner
            if user.role != 'shop':
                return Response(
                    {"success": False, "message": "This account is not registered as a shop"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if shop exists and is verified
            try:
                shop = Shop.objects.get(user=user)
                if not shop.is_email_verified:
                    return Response(
                        {"success": False, "message": "Please verify your email before logging in"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except Shop.DoesNotExist:
                return Response(
                    {"success": False, "message": "Shop profile not found"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Proceed with authentication
            response = super().post(request, *args, **kwargs)
            token = response.data

            access_token = token["access"]
            refresh_token = token["refresh"]

            res = Response(status=status.HTTP_200_OK)

            res.data = {
                "success": True,
                "message": "Shop login successfully",
                "shopDetails": {
                    "id": shop.id,
                    "name": shop.name,
                    "email": shop.email,
                    "owner_name": shop.owner_name,
                    "is_approved": shop.is_approved,
                    "is_email_verified": shop.is_email_verified
                },
                "userDetails": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "role": user.role
                }
            }

            # FIXED: Set cookies with root path for all shop routes
            res.set_cookie(
                key="access_token",
                value=access_token,
                httponly=True,
                secure=False,  # Set to True in production with HTTPS
                samesite="Lax",  # Changed from "None" to "Lax" for local development
                path='/',  # FIXED: Use root path instead of specific path
                max_age=3600
            )

            res.set_cookie(
                key="refresh_token",
                value=refresh_token,
                httponly=True,
                secure=False,  # Set to True in production with HTTPS
                samesite="Lax",  # Changed from "None" to "Lax" for local development
                path='/',  # FIXED: Use root path instead of specific path
                max_age=86400  # 24 hours for refresh token
            )
            return res

        except AuthenticationFailed:
            return Response(
                {"success": False, "message": "Invalid Credentials"},
                status=status.HTTP_400_BAD_REQUEST
            )

        except Exception as e:
            return Response(
                {"success": False, "message": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CustomShopTokenRefreshView(TokenRefreshView):
    permission_classes = [permissions.AllowAny]
    
    def post(self, request, *args, **kwargs):
        try:
            refresh_token = request.COOKIES.get("refresh_token")
            
            if not refresh_token:
                return Response(
                    {"success": False, "message": "Refresh token not found in cookies"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Use the parent class's post method to handle token refresh properly
            # This ensures proper token rotation and blacklisting
            original_data = request.data
            request._full_data = {"refresh": refresh_token}
            
            try:
                # Call parent's post method which handles token rotation
                response = super().post(request, *args, **kwargs)
                
                if response.status_code == 200:
                    token_data = response.data
                    access_token = token_data.get("access")
                    new_refresh_token = token_data.get("refresh")  # This might be None if rotation is disabled
                    
                    # Create success response
                    res = Response(
                        {"success": True, "message": "Access token refreshed"},
                        status=status.HTTP_200_OK
                    )

                    # Set the new access token cookie
                    res.set_cookie(
                        key="access_token",
                        value=access_token,
                        httponly=True,
                        secure=False,  # Set to True in production
                        samesite="Lax",
                        path='/',
                        max_age=3600
                    )
                    
                    # Set new refresh token cookie if token rotation is enabled
                    if new_refresh_token:
                        res.set_cookie(
                            key="refresh_token",
                            value=new_refresh_token,
                            httponly=True,
                            secure=False,  # Set to True in production
                            samesite="Lax",
                            path='/',
                            max_age=86400  # 24 hours
                        )

                    return res
                else:
                    return Response(
                        {"success": False, "message": "Invalid or expired refresh token"},
                        status=status.HTTP_401_UNAUTHORIZED
                    )
                    
            except Exception as token_error:
                # Handle specific token errors
                error_message = str(token_error)
                if "blacklisted" in error_message.lower():
                    return Response(
                        {"success": False, "message": "Session expired. Please login again."},
                        status=status.HTTP_401_UNAUTHORIZED
                    )
                else:
                    return Response(
                        {"success": False, "message": "Invalid or expired refresh token"},
                        status=status.HTTP_401_UNAUTHORIZED
                    )
            finally:
                # Restore original request data
                request._full_data = original_data
            
        except Exception as e:
            return Response(
                {"success": False, "message": f"Token refresh failed: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )


class ShopLogoutView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            # Get tokens from cookies
            access_token = request.COOKIES.get('access_token')
            refresh_token = request.COOKIES.get('refresh_token')
            
            # Blacklist the refresh token if it exists
            if refresh_token:
                try:
                    token = RefreshToken(refresh_token)
                    token.blacklist()
                    print(f"Refresh token blacklisted successfully")
                except TokenError as e:
                    # Token might already be blacklisted or invalid
                    print(f"Token blacklist error: {str(e)}")
            
            # Create response
            res = Response(status=status.HTTP_200_OK)
            res.data = {"success": True, "message": "Logged out successfully"}
            
            # Clear cookies with consistent paths
            res.delete_cookie('access_token', path='/')
            res.delete_cookie('refresh_token', path='/')
            
            return res
            
        except Exception as e:
            return Response(
                {"success": False, "message": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

logger = logging.getLogger(__name__)

class ShopProfileView(APIView):
    """View to retrieve the authenticated shop's profile information."""
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            # Check if user is authenticated
            if not request.user.is_authenticated:
                logger.warning("Unauthenticated user trying to access shop profile")
                return Response(
                    {"success": False, "error": "Authentication required"}, 
                    status=status.HTTP_401_UNAUTHORIZED
                )
            
            # Check if user has shop role
            if getattr(request.user, 'role', None) != 'shop':
                logger.warning(f"User {request.user.id} with role {getattr(request.user, 'role', 'None')} trying to access shop profile")
                return Response(
                    {"success": False, "error": "Only shop owners can access this endpoint"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            try:
                shop = Shop.objects.get(user=request.user)
                logger.debug(f"Found shop: {shop.name} for user: {request.user.username}")
                
                serializer = ShopSerializer(shop)
                return Response({
                    "success": True,
                    "data": serializer.data
                })
            except Shop.DoesNotExist:
                logger.error(f"Shop not found for user: {request.user.id}")
                return Response(
                    {"success": False, "error": "Shop profile not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
                
        except Exception as e:
            logger.error(f"Unexpected error in ShopProfileView: {str(e)}")
            return Response(
                {"success": False, "error": "Internal server error"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ShopUpdateView(APIView):
    """View to update shop details and handle images."""
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            # Check if user has shop role
            if getattr(request.user, 'role', None) != 'shop':
                logger.warning(f"User {request.user.id} with role {getattr(request.user, 'role', 'None')} trying to update shop profile")
                return Response(
                    {"success": False, "error": "Only shop owners can access this endpoint"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            shop = get_object_or_404(Shop, user=request.user)
            
            # Update basic shop details
            shop.name = request.data.get('name', shop.name)
            shop.phone = request.data.get('phone', shop.phone)
            shop.owner_name = request.data.get('owner_name', shop.owner_name)
            shop.address = request.data.get('address', shop.address)
            shop.description = request.data.get('description', shop.description)
            shop.save()
            
            # Handle images if provided
            images_data = request.data.get('images', [])
            if images_data:
                # Clear existing images
                ShopImage.objects.filter(shop=shop).delete()
                
                # Add new images
                for index, image_data in enumerate(images_data):
                    ShopImage.objects.create(
                        shop=shop,
                        image_url=image_data.get('url'),
                        public_id=image_data.get('public_id'),
                        width=image_data.get('width'),
                        height=image_data.get('height'),
                        is_primary=(index == 0),  # First image is primary
                        order=index
                    )
            
            # Return updated shop data
            serializer = ShopSerializer(shop)
            return Response({
                'success': True,
                'data': serializer.data,
                'message': 'Shop profile updated successfully'
            })
            
        except Shop.DoesNotExist:
            return Response(
                {"success": False, "error": "Shop profile not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Unexpected error in ShopUpdateView: {str(e)}")
            return Response(
                {"success": False, "error": str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ShopImageAddView(APIView):
    """View to add a new image to shop."""
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            # Check if user has shop role
            if getattr(request.user, 'role', None) != 'shop':
                return Response(
                    {"success": False, "error": "Only shop owners can access this endpoint"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            shop = get_object_or_404(Shop, user=request.user)
            
            image_data = request.data
            order = ShopImage.objects.filter(shop=shop).count()
            is_primary = order == 0  # First image is primary
            
            shop_image = ShopImage.objects.create(
                shop=shop,
                image_url=image_data.get('url'),
                public_id=image_data.get('public_id'),
                width=image_data.get('width'),
                height=image_data.get('height'),
                is_primary=is_primary,
                order=order
            )
            
            serializer = ShopImageSerializer(shop_image)
            return Response({
                'success': True,
                'data': serializer.data,
                'message': 'Image added successfully'
            })
            
        except Exception as e:
            logger.error(f"Error in ShopImageAddView: {str(e)}")
            return Response(
                {"success": False, "error": str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ShopImageRemoveView(APIView):
    """View to remove a shop image."""
    
    permission_classes = [IsAuthenticated]
    
    def delete(self, request, image_id):
        try:
            # Check if user has shop role
            if getattr(request.user, 'role', None) != 'shop':
                return Response(
                    {"success": False, "error": "Only shop owners can access this endpoint"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            shop = get_object_or_404(Shop, user=request.user)
            image = get_object_or_404(ShopImage, id=image_id, shop=shop)
            
            was_primary = image.is_primary
            image.delete()
            
            # If deleted image was primary, make the first remaining image primary
            if was_primary:
                first_image = ShopImage.objects.filter(shop=shop).first()
                if first_image:
                    first_image.is_primary = True
                    first_image.save()
            
            return Response({
                'success': True,
                'message': 'Image removed successfully'
            })
            
        except Exception as e:
            logger.error(f"Error in ShopImageRemoveView: {str(e)}")
            return Response(
                {"success": False, "error": str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ShopImageSetPrimaryView(APIView):
    """View to set an image as primary."""
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request, image_id):
        try:
            # Check if user has shop role
            if getattr(request.user, 'role', None) != 'shop':
                return Response(
                    {"success": False, "error": "Only shop owners can access this endpoint"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            shop = get_object_or_404(Shop, user=request.user)
            image = get_object_or_404(ShopImage, id=image_id, shop=shop)
            
            # Set all images as non-primary
            ShopImage.objects.filter(shop=shop).update(is_primary=False)
            
            # Set this image as primary
            image.is_primary = True
            image.save()
            
            return Response({
                'success': True,
                'message': 'Primary image updated successfully'
            })
            
        except Exception as e:
            logger.error(f"Error in ShopImageSetPrimaryView: {str(e)}")
            return Response(
                {"success": False, "error": str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class PublicShopDetailView(APIView):
    """Public view to get shop details for users."""
    
    def get(self, request, shop_id):
        try:
            shop = get_object_or_404(Shop, id=shop_id, is_approved=True)
            serializer = ShopSerializer(shop)
            return Response({
                'success': True,
                'data': serializer.data
            })
        except Shop.DoesNotExist:
            return Response(
                {"success": False, "error": "Shop not found or not approved"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error in PublicShopDetailView: {str(e)}")
            return Response(
                {"success": False, "error": "Internal server error"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class PublicShopListView(APIView):
    """Public view to get all approved shops."""
    
    def get(self, request):
        try:
            shops = Shop.objects.filter(is_approved=True)
            serializer = ShopSerializer(shops, many=True)
            return Response({
                'success': True,
                'data': serializer.data
            })
        except Exception as e:
            logger.error(f"Error in PublicShopListView: {str(e)}")
            return Response(
                {"success": False, "error": str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# class ShopDashboardStatsView(APIView):
#     """View to retrieve shop dashboard statistics."""
    
#     permission_classes = [IsAuthenticated]
    
#     def get(self, request):
#         try:
#             shop = Shop.objects.get(user=request.user)
            
#             # Check if shop is approved
#             if not shop.is_approved:
#                 return Response(
#                     {"error": "Shop is not approved yet"}, 
#                     status=status.HTTP_403_FORBIDDEN
#                 )
            
#             # Get today's date to filter appointments
#             from datetime import datetime, time
#             today = datetime.now().date()
#             today_start = datetime.combine(today, time.min)
#             today_end = datetime.combine(today, time.max)
            
#             # Calculate statistics
#             appointments_today = Appointment.objects.filter(
#                 shop=shop, 
#                 start_time__range=(today_start, today_end)
#             ).count()
            
#             # Calculating total sales (this is a simplified example)
#             total_sales = Appointment.objects.filter(
#                 shop=shop, 
#                 status='completed'
#             ).aggregate(total=Sum('price'))['total'] or 0
            
#             # Count unique customers
#             total_customers = CustomUser.objects.filter(appointments__shop=shop).distinct().count()
            
#             # Count services/products offered by this shop
#             total_services = Service.objects.filter(shop=shop).count()
            
#             stats = {
#                 'appointments_today': appointments_today,
#                 'total_sales': total_sales,
#                 'total_customers': total_customers,
#                 'products': total_services  # Called 'products' to match frontend expectation
#             }
            
#             return Response(stats)
            
#         except Shop.DoesNotExist:
#             return Response(
#                 {"error": "Shop profile not found"}, 
#                 status=status.HTTP_404_NOT_FOUND
#             )


# Shop Data Views
# class RecentAppointmentsView(APIView):
#     """View to retrieve recent appointments for a shop."""
    
#     permission_classes = [IsAuthenticated]
    
#     def get(self, request):
#         try:
#             shop = Shop.objects.get(user=request.user)
#             # Get the most recent appointments (e.g., limited to 5)
#             appointments = Appointment.objects.filter(shop=shop).order_by('-start_time')[:5]
#             serializer = AppointmentSerializer(appointments, many=True)
#             return Response(serializer.data)
#         except Shop.DoesNotExist:
#             return Response(
#                 {"error": "Shop profile not found"}, 
#                 status=status.HTTP_404_NOT_FOUND
#             )


class ShopNotificationsView(APIView):
    """View to retrieve shop notifications."""
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            shop = Shop.objects.get(user=request.user)
            # Get unread notifications first, then read ones, most recent first
            notifications = Notification.objects.filter(shop=shop).order_by('-read', '-created_at')[:10]
            serializer = NotificationSerializer(notifications, many=True)
            return Response(serializer.data)
        except Shop.DoesNotExist:
            return Response(
                {"error": "Shop profile not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )


class BusinessHoursView(APIView):
    """API view for business hours operations using DRF."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get all business hours for the authenticated user's shop."""
        try:
            shop = request.user.shop
        except Shop.DoesNotExist:
            return Response(
                {'error': 'No shop associated with this user'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get business hours for this specific shop
        hours = BusinessHours.objects.filter(shop=shop)
        serializer = BusinessHoursSerializer(hours, many=True)
        return Response(serializer.data)


class BusinessHoursUpdateView(APIView):
    """API view for updating business hours using DRF."""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """Update business hours."""
        try:
            # Get the shop associated with the authenticated user
            try:
                shop = request.user.shop
            except Shop.DoesNotExist:
                return Response(
                    {'error': 'No shop associated with this user'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            business_hours = request.data.get('business_hours', [])
            updated_hours = []
            
            for hour_data in business_hours:
                day_of_week = hour_data.get('day_of_week')
                is_closed = hour_data.get('is_closed', False)
                
                if day_of_week is None:
                    continue
                
                # Create defaults dictionary
                defaults = {'is_closed': is_closed}
                
                # Always set default times - even for closed days
                # This ensures NOT NULL constraints are satisfied
                opening_time = hour_data.get('opening_time', '09:00')
                closing_time = hour_data.get('closing_time', '17:00')
                
                try:
                    # Parse time strings - assuming format like "09:00"
                    from datetime import time
                    
                    if isinstance(opening_time, str) and opening_time:
                        hour, minute = map(int, opening_time.split(':'))
                        defaults['opening_time'] = time(hour=hour, minute=minute)
                    else:
                        # Default opening time if not provided or invalid
                        defaults['opening_time'] = time(hour=9, minute=0)
                    
                    if isinstance(closing_time, str) and closing_time:
                        hour, minute = map(int, closing_time.split(':'))
                        defaults['closing_time'] = time(hour=hour, minute=minute)
                    else:
                        # Default closing time if not provided or invalid
                        defaults['closing_time'] = time(hour=17, minute=0)
                except Exception as e:
                    return Response({'error': f'Invalid time format: {str(e)}'}, 
                                   status=status.HTTP_400_BAD_REQUEST)
                
                # Create or update the business hours - NOW INCLUDING THE SHOP
                hours, created = BusinessHours.objects.update_or_create(
                    shop=shop,  # Include the shop in the lookup
                    day_of_week=day_of_week,
                    defaults=defaults
                )
                
                serializer = BusinessHoursSerializer(hours)
                updated_hours.append(serializer.data)
            
            return Response({
                'success': True,
                'message': 'Business hours updated successfully',
                'business_hours': updated_hours
            })
        
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# class AvailableSlotsView(APIView):
#     """API view for available time slots using DRF."""
#     permission_classes = [IsAuthenticated]
    
#     def get(self, request):
#         """Get available time slots for a specific date."""
#         date_str = request.query_params.get('date')
#         barber_id = request.query_params.get('barber_id')
#         service_id = request.query_params.get('service_id')
        
#         if not date_str:
#             return Response({'error': 'Date parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        
#         try:
#             date = parse(date_str).date()
#         except ValueError:
#             return Response({'error': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)
        
#         # Get selected barber if provided
#         barber = None
#         if barber_id:
#             try:
#                 barber = Barber.objects.get(id=barber_id)
#             except Barber.DoesNotExist:
#                 return Response({'error': 'Barber not found'}, status=status.HTTP_404_NOT_FOUND)
        
#         # Get service duration if provided
#         slot_duration = 30  # Default 30 min
#         if service_id:
#             try:
#                 service = Service.objects.get(id=service_id)
#                 slot_duration = service.duration.total_seconds() // 60  # Convert to minutes
#             except Service.DoesNotExist:
#                 return Response({'error': 'Service not found'}, status=status.HTTP_404_NOT_FOUND)
        
#         # Get available slots
#         slots = get_available_slots(date, barber, slot_duration)
        
#         # Format slots for response
#         formatted_slots = []
#         for start, end in slots:
#             formatted_slots.append({
#                 'start_time': start.strftime('%Y-%m-%d %H:%M'),
#                 'end_time': end.strftime('%Y-%m-%d %H:%M'),
#                 'display': f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}"
#             })
        
#         return Response({
#             'date': date.strftime('%Y-%m-%d'),
#             'available_slots': formatted_slots
#         })


# class AppointmentCreateView(APIView):
#     """API view for creating appointments using DRF."""
#     permission_classes = [IsAuthenticated]
    
#     def post(self, request):
#         """Create a new appointment."""
#         try:
#             serializer = AppointmentSerializer(data=request.data)
#             if serializer.is_valid():
#                 # Extract data
#                 customer_id = serializer.validated_data.get('customer_id')
#                 barber_id = serializer.validated_data.get('barber_id')
#                 service_id = serializer.validated_data.get('service_id')
#                 start_time = serializer.validated_data.get('start_time')
                
#                 # Get the service for duration
#                 service = Service.objects.get(id=service_id)
#                 end_time = start_time + service.duration
                
#                 # Check if the slot is still available
#                 date = start_time.date()
#                 available_slots = get_available_slots(date, Barber.objects.get(id=barber_id))
                
#                 slot_available = False
#                 for slot_start, slot_end in available_slots:
#                     if slot_start == start_time:
#                         slot_available = True
#                         break
                
#                 if not slot_available:
#                     return Response({'error': 'This time slot is no longer available'}, 
#                                     status=status.HTTP_400_BAD_REQUEST)
                
#                 # Create the appointment
#                 appointment = serializer.save(end_time=end_time)
                
#                 return Response({
#                     'success': True,
#                     'appointment_id': appointment.id,
#                     'start_time': appointment.start_time.strftime('%Y-%m-%d %H:%M'),
#                     'end_time': appointment.end_time.strftime('%Y-%m-%d %H:%M')
#                 }, status=status.HTTP_201_CREATED)
            
#             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
#         except Exception as e:
#             return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class SpecialClosingDayView(APIView):
    """API view for managing special closing days using DRF."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get all special closing days for the authenticated shop owner."""
        try:
            # Use 'user' instead of 'owner'
            shop = Shop.objects.get(user=request.user)
            closing_days = SpecialClosingDay.objects.filter(shop=shop).order_by('date')
            serializer = SpecialClosingDaySerializer(closing_days, many=True)
            return Response(serializer.data)
        except Shop.DoesNotExist:
            return Response({'error': 'Shop not found'}, status=status.HTTP_404_NOT_FOUND)
    
    def post(self, request):
        """Add a special closing day and cancel existing bookings."""
        try:
            # Use 'user' instead of 'owner'
            shop = Shop.objects.get(user=request.user)
            
            with transaction.atomic():
                # Add shop to the request data
                data = request.data.copy()
                data['shop'] = shop.id
                
                serializer = SpecialClosingDaySerializer(data=data)
                if serializer.is_valid():
                    closing_day = serializer.save()
                    
                    # Find all bookings on this date FOR THIS SHOP ONLY
                    affected_bookings = Booking.objects.filter(
                        shop=shop,  # Filter by shop
                        appointment_date=closing_day.date,
                        booking_status__in=['pending', 'confirmed']
                    )
                    
                    cancelled_bookings = []
                    refunded_amount = 0
                    
                    # Cancel each booking and process refunds
                    for booking in affected_bookings:
                        if booking.cancel_with_refund(f"Shop closed - {closing_day.reason}"):
                            cancelled_bookings.append({
                                'booking_id': booking.id,
                                'user': booking.user.username,
                                'amount_refunded': float(booking.total_amount)
                            })
                            refunded_amount += booking.total_amount
                    
                    # Send notifications (same logic as before)
                    for booking in affected_bookings:
                        if booking.booking_status == 'cancelled':
                            try:
                                user = booking.user
                                shop_owner = request.user
                                
                                # Check if user exists and is different from shop owner
                                if user and user.id != shop_owner.id:
                                    # Create notification message
                                    notification_message = f"Your booking for {booking.shop.name} on {closing_day.date.strftime('%Y-%m-%d')} has been cancelled - {closing_day.reason}. Refund: ₹{booking.total_amount}"
                                    
                                    # Create database notification FIRST
                                    logger.info(f"Creating database notification for user {user.id}")
                                    
                                    notification = Notification.objects.create(
                                        sender=shop_owner,  # Shop owner is the sender
                                        receiver=user,  # Booking user is the receiver
                                        message=notification_message,
                                        is_read=False
                                    )
                                    logger.info(f"Database notification created with ID: {notification.id}")
                                    
                                    # Then send real-time notification if user is online
                                    ONLINE_USERS = f'chat:online_users'
                                    curr_users = cache.get(ONLINE_USERS, [])
                                    logger.info(f"Current online users: {curr_users}")
                                    
                                    # Check if user is online
                                    is_user_online = user.id in [online_user["id"] for online_user in curr_users] if curr_users else False
                                    logger.info(f"User {user.id} online status: {is_user_online}")
                                    
                                    if is_user_online:
                                        logger.info("Sending real-time notification to user")
                                        channel_layer = get_channel_layer()
                                        if channel_layer:
                                            data = {
                                                'type': 'notification',
                                                'message': {
                                                    'sender': shop_owner.username,
                                                    'content': notification_message,
                                                    'booking_id': booking.id,
                                                    'shop_name': booking.shop.name,
                                                    'appointment_date': closing_day.date.strftime('%Y-%m-%d'),
                                                    'appointment_time': booking.appointment_time.strftime('%H:%M'),
                                                    'cancellation_reason': closing_day.reason,
                                                    'refund_amount': float(booking.total_amount),
                                                    'notification_type': 'booking_cancellation'
                                                }
                                            }
                                            
                                            group_name = f'user_{user.id}'
                                            async_to_sync(channel_layer.group_send)(group_name, data)
                                            logger.info(f"Real-time notification sent to user {user.id}")
                                    else:
                                        logger.info("User is offline, notification stored in database only")
                                else:
                                    logger.warning("User is None or same as shop owner - skipping notification")
                                            
                            except Exception as notification_error:
                                logger.error(f"Notification failed for booking {booking.id}: {str(notification_error)}")
                                logger.error(f"Notification error type: {type(notification_error)}")
                                # Continue with other bookings - don't fail the entire process
                                continue
                    
                    return Response({
                        'success': True,
                        'message': 'Special closing day added successfully',
                        'closing_day': {
                            'id': closing_day.id,
                            'date': closing_day.date.strftime('%Y-%m-%d'),
                            'reason': closing_day.reason
                        },
                        'cancelled_bookings_count': len(cancelled_bookings),
                        'total_refunded_amount': float(refunded_amount),
                        'cancelled_bookings': cancelled_bookings
                    }, status=status.HTTP_201_CREATED)
                
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except Shop.DoesNotExist:
            return Response({'error': 'Shop not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error adding special closing day: {str(e)}")
            return Response({
                'error': f'Failed to add special closing day: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)


class SpecialClosingDayDetailView(APIView):
    """API view for managing a specific special closing day using DRF."""
    permission_classes = [IsAuthenticated]
    
    def delete(self, request, id):
        """Remove a special closing day."""
        try:
            # Use 'user' instead of 'owner'
            shop = Shop.objects.get(user=request.user)
            
            # Ensure the closing day belongs to this shop
            closing_day = get_object_or_404(SpecialClosingDay, id=id, shop=shop)
            closing_day.delete()
            
            return Response({
                'success': True,
                'message': 'Special closing day removed successfully'
            }, status=status.HTTP_200_OK)
            
        except Shop.DoesNotExist:
            return Response({'error': 'Shop not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

class ServiceListView(generics.ListAPIView):
    """
    API view for listing all services belonging to the authenticated shop.
    """
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Return services belonging to the authenticated user's shop.
        Assumes the authenticated user has a related shop.
        """
        try:
            # If your User model has a direct relationship to Shop
            if hasattr(self.request.user, 'shop'):
                return Service.objects.filter(shop=self.request.user.shop)
            # If the User IS the Shop (your current setup)
            else:
                return Service.objects.filter(shop=self.request.user)
        except AttributeError:
            return Service.objects.none()


class ServiceCreateView(generics.CreateAPIView):
    """
    API view for creating a new service.
    """
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        logger.info(f"Creating service for user: {self.request.user}")
        
        # Automatically use the authenticated user as the shop
        try:
            # If your User model has a direct relationship to Shop
            if hasattr(self.request.user, 'shop'):
                shop = self.request.user.shop
                logger.info(f"Using user.shop: {shop}")
            # If the User IS the Shop (your current setup based on other views)
            else:
                shop = self.request.user
                logger.info(f"Using user as shop: {shop}")
                
            serializer.save(shop=shop)
            logger.info("Service created successfully")
        except AttributeError as e:
            logger.error(f"AttributeError in perform_create: {e}")
            raise ValidationError({"error": "Unable to determine shop for authenticated user"})
        except Exception as e:
            logger.error(f"Exception in perform_create: {e}")
            raise ValidationError({"error": "Failed to create service"})


class ServiceUpdateView(generics.RetrieveUpdateAPIView):
    """
    API view for updating a service.
    """
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    
    def get_queryset(self):
        logger.info(f"Getting queryset for user: {self.request.user}")
        
        # Get services for the authenticated shop
        try:
            if hasattr(self.request.user, 'shop'):
                shop = self.request.user.shop
            else:
                shop = self.request.user
            
            queryset = Service.objects.filter(shop=shop)
            logger.info(f"Queryset for shop {shop}: {queryset.count()} services")
            return queryset
        except AttributeError as e:
            logger.error(f"AttributeError in get_queryset: {e}")
            return Service.objects.none()
    
    def get_object(self):
        obj = get_object_or_404(self.get_queryset(), id=self.kwargs['id'])
        logger.info(f"Retrieved object: {obj}")
        return obj
    
    def perform_update(self, serializer):
        logger.info(f"Updating service for user: {self.request.user}")
        
        # Ensure the shop remains the same during update
        try:
            if hasattr(self.request.user, 'shop'):
                shop = self.request.user.shop
            else:
                shop = self.request.user
                
            serializer.save(shop=shop)
            logger.info("Service updated successfully")
        except AttributeError as e:
            logger.error(f"AttributeError in perform_update: {e}")
            raise ValidationError({"error": "Unable to determine shop for authenticated user"})
        except Exception as e:
            logger.error(f"Exception in perform_update: {e}")
            raise ValidationError({"error": "Failed to update service"})


# DEBUG: Add a simple test view to check what's happening
class ServiceDebugView(generics.GenericAPIView):
    """
    Debug view to test validation manually.
    """
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated]
    
    def post(self, request, *args, **kwargs):
        logger.info(f"Debug view called with data: {request.data}")
        
        # Test the serializer validation manually
        serializer = self.get_serializer(data=request.data)
        
        try:
            serializer.is_valid(raise_exception=True)
            logger.info("Validation passed")
            return Response({"message": "Validation passed", "data": serializer.validated_data})
        except serializers.ValidationError as e:
            logger.error(f"Validation failed: {e}")
            return Response({"error": "Validation failed", "details": e.detail}, status=400)



class ServiceDeleteView(generics.DestroyAPIView):
    """
    API view for deleting an existing service.
    Only allows deleting services belonging to the authenticated shop.
    """
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'

    def get_queryset(self):
        """
        Return services belonging to the authenticated user (shop) only.
        """
        try:
            shop = self.request.user.shop
            return Service.objects.filter(shop=shop)
        except Shop.DoesNotExist:
            return Service.objects.none()
        
    def get_object(self):
        """
        Override to ensure proper permission checking.
        """
        obj = get_object_or_404(self.get_queryset(), id=self.kwargs['id'])
        self.check_object_permissions(self.request, obj)
        return obj

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Service deleted successfully"}, status=status.HTTP_200_OK)
    

class ShopForgotPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = ShopForgotPasswordSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            
            # Get shop information for personalized email
            try:
                shop = Shop.objects.get(user=user)
                shop_name = shop.name
                owner_name = shop.owner_name
            except Shop.DoesNotExist:
                shop_name = "Your Shop"  # Fallback
                owner_name = user.get_full_name() or user.username
            
            # Send forgot password OTP email asynchronously
            send_shop_forgot_password_otp_task.delay(
                email=user.email,
                otp=user.otp,
                shop_name=shop_name,
                owner_name=owner_name
            )
            
            logger.info(f"Shop forgot password OTP email task queued for {user.email}")
            
            return Response({
                "message": "Password reset code has been sent to your email"
            }, status=status.HTTP_200_OK)
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    

class ShopVerifyForgotPasswordOTPView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = ShopVerifyForgotPasswordOTPSerializer(data=request.data)
        if serializer.is_valid():
            return Response({"message": "OTP verified successfully"}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ShopResetPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = ShopResetPasswordSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response({"message": "Password reset successfully"}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)











class UserBookingFeedbackListView(APIView):
    """
    Get all feedback submitted by the authenticated user
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        """Get all feedback by the user"""
        try:
            feedbacks = BookingFeedback.objects.filter(
                user=request.user
            ).select_related('booking', 'shop').order_by('-created_at')
            
            serializer = BookingFeedbackSerializer(feedbacks, many=True)
            
            return Response({
                'feedbacks': serializer.data,
                'count': feedbacks.count()
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': 'Failed to fetch feedbacks'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ShopFeedbackListView(APIView):
    """
    Get all feedback for a specific shop (for shop owners or public view)
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, shop_id, *args, **kwargs):
        """Get all feedback for a shop"""
        try:
            # You might want to add additional permission checks here
            # For example, only allow shop owners to see their feedback
            
            feedbacks = BookingFeedback.objects.filter(
                shop_id=shop_id
            ).select_related('user', 'booking').order_by('-created_at')
            
            # Optional: Hide user sensitive information for public view
            serializer = BookingFeedbackSerializer(feedbacks, many=True)
            
            # Calculate average ratings
            if feedbacks.exists():
                avg_rating = feedbacks.aggregate(
                    avg_overall=models.Avg('rating'),
                    avg_service=models.Avg('service_quality'),
                    avg_staff=models.Avg('staff_behavior'),
                    avg_cleanliness=models.Avg('cleanliness'),
                    avg_value=models.Avg('value_for_money')
                )
            else:
                avg_rating = {
                    'avg_overall': 0,
                    'avg_service': 0,
                    'avg_staff': 0,
                    'avg_cleanliness': 0,
                    'avg_value': 0
                }
            
            return Response({
                'feedbacks': serializer.data,
                'count': feedbacks.count(),
                'averages': avg_rating
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': 'Failed to fetch shop feedback'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from django.db.models import Avg, Count
from rest_framework.decorators import api_view


@api_view(['GET'])
def get_shop_rating_summary(request, shop_id):
    """
    Get rating summary for a specific shop
    """
    try:
        shop = Shop.objects.get(id=shop_id)
    except Shop.DoesNotExist:
        return Response(
            {'error': 'Shop not found'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Get all feedback for this shop
    feedback_queryset = BookingFeedback.objects.filter(shop=shop)
    
    # Calculate ratings
    rating_aggregates = feedback_queryset.aggregate(
        average_rating=Avg('rating'),
        total_reviews=Count('id'),
        avg_service_quality=Avg('service_quality'),
        avg_staff_behavior=Avg('staff_behavior'),
        avg_cleanliness=Avg('cleanliness'),
        avg_value_for_money=Avg('value_for_money')
    )
    
    # Prepare response data - simplified for just overall rating
    response_data = {
        'shop_id': shop.id,
        'shop_name': shop.name,
        'average_rating': round(rating_aggregates['average_rating'] or 0, 1),
        'total_reviews': rating_aggregates['total_reviews'] or 0
    }
    
    return Response(response_data, status=status.HTTP_200_OK)













from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, Count, Avg, Q, F, Min
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import HttpResponse
import json
from collections import defaultdict




class SalesChartView(APIView):
    """Get sales data for chart visualization"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        period = int(request.GET.get('period', 7))
        shop = request.user.shop
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period)
        
        # Get daily sales data
        bookings = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date,
            booking_status__in=['completed', 'confirmed']
        ).values('appointment_date').annotate(
            revenue=Sum('total_amount'),
            bookings_count=Count('id')
        ).order_by('appointment_date')
        
        # Format data for chart
        sales_data = []
        for booking in bookings:
            sales_data.append({
                'date': booking['appointment_date'].strftime('%Y-%m-%d'),
                'revenue': float(booking['revenue'] or 0),
                'bookings': booking['bookings_count']
            })
        
        return Response({
            'sales_data': sales_data,
            'period': period
        })


class MostBookedServicesView(APIView):
    """Get most booked services data"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        period = int(request.GET.get('period', 7))
        shop = request.user.shop
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period)
        
        # Get service booking data
        service_stats = []
        
        # Get bookings for the period
        bookings = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date,
            booking_status__in=['completed', 'confirmed']
        ).prefetch_related('services')
        
        # Count services
        service_counts = defaultdict(lambda: {'bookings': 0, 'revenue': 0})
        
        for booking in bookings:
            for service in booking.services.all():
                service_counts[service.name]['bookings'] += 1
                service_counts[service.name]['revenue'] += float(booking.total_amount) / booking.services.count()
        
        # Convert to list and sort
        services = []
        for name, stats in service_counts.items():
            services.append({
                'name': name,
                'bookings': stats['bookings'],
                'revenue': stats['revenue']
            })
        
        services.sort(key=lambda x: x['bookings'], reverse=True)
        
        return Response({
            'services': services[:10],  # Top 10
            'period': period
        })


class RevenueStatsView(APIView):
    """Get revenue and booking statistics"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        period = int(request.GET.get('period', 7))
        shop = request.user.shop
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period)
        
        # Current period stats
        current_bookings = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date
        )
        
        current_stats = current_bookings.aggregate(
            total_revenue=Sum('total_amount', filter=Q(booking_status='completed')),
            total_bookings=Count('id'),
            completed_bookings=Count('id', filter=Q(booking_status='completed')),
            pending_bookings=Count('id', filter=Q(booking_status='pending')),
            cancelled_bookings=Count('id', filter=Q(booking_status='cancelled')),
            avg_order_value=Avg('total_amount', filter=Q(booking_status='completed'))
        )
        
        # Previous period for comparison
        prev_start_date = start_date - timedelta(days=period)
        prev_end_date = start_date - timedelta(days=1)
        
        prev_bookings = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=prev_start_date,
            appointment_date__lte=prev_end_date
        )
        
        prev_stats = prev_bookings.aggregate(
            total_revenue=Sum('total_amount', filter=Q(booking_status='completed'))
        )
        
        # Calculate growth rate
        current_revenue = float(current_stats['total_revenue'] or 0)
        prev_revenue = float(prev_stats['total_revenue'] or 0)
        
        growth_rate = 0
        if prev_revenue > 0:
            growth_rate = ((current_revenue - prev_revenue) / prev_revenue) * 100
        
        # Calculate completion rate
        total_bookings = current_stats['total_bookings'] or 0
        completed_bookings = current_stats['completed_bookings'] or 0
        completion_rate = (completed_bookings / total_bookings * 100) if total_bookings > 0 else 0
        
        return Response({
            'totalRevenue': current_revenue,
            'totalBookings': total_bookings,
            'completedBookings': completed_bookings,
            'pendingBookings': current_stats['pending_bookings'] or 0,
            'cancelledBookings': current_stats['cancelled_bookings'] or 0,
            'avgOrderValue': float(current_stats['avg_order_value'] or 0),
            'completionRate': round(completion_rate, 1),
            'growthRate': round(growth_rate, 1),
            'period': period
        })


class ServicePerformanceView(APIView):
    """Get detailed service performance data"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        period = int(request.GET.get('period', 7))
        shop = request.user.shop
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period)
        
        # Get services for the shop
        services = Service.objects.filter(shop=shop, is_active=True)
        
        service_performance = []
        
        for service in services:
            # Current period bookings
            current_bookings = Booking.objects.filter(
                shop=shop,
                services=service,
                appointment_date__gte=start_date,
                appointment_date__lte=end_date,
                booking_status='completed'
            )
            
            current_stats = current_bookings.aggregate(
                bookings=Count('id'),
                revenue=Sum('total_amount')
            )
            
            # Previous period for comparison
            prev_start_date = start_date - timedelta(days=period)
            prev_end_date = start_date - timedelta(days=1)
            
            prev_bookings = Booking.objects.filter(
                shop=shop,
                services=service,
                appointment_date__gte=prev_start_date,
                appointment_date__lte=prev_end_date,
                booking_status='completed'
            )
            
            prev_stats = prev_bookings.aggregate(
                bookings=Count('id')
            )
            
            # Calculate growth
            current_booking_count = current_stats['bookings'] or 0
            prev_booking_count = prev_stats['bookings'] or 0
            
            growth = 0
            if prev_booking_count > 0:
                growth = ((current_booking_count - prev_booking_count) / prev_booking_count) * 100
            
            # Calculate average price
            revenue = float(current_stats['revenue'] or 0)
            avg_price = revenue / current_booking_count if current_booking_count > 0 else float(service.price)
            
            if current_booking_count > 0:  # Only include services with bookings
                service_performance.append({
                    'name': service.name,
                    'bookings': current_booking_count,
                    'revenue': revenue,
                    'avgPrice': avg_price,
                    'growth': round(growth, 1)
                })
        
        # Sort by bookings count
        service_performance.sort(key=lambda x: x['bookings'], reverse=True)
        
        return Response({
            'services': service_performance,
            'period': period
        })


class PaymentMethodStatsView(APIView):
    """Get payment method statistics"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        period = int(request.GET.get('period', 7))
        shop = request.user.shop
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period)
        
        # Get payment method stats
        payment_stats = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date,
            booking_status='completed'
        ).values('payment_method').annotate(
            total_amount=Sum('total_amount'),
            count=Count('id')
        ).order_by('-total_amount')
        
        # Format data for pie chart
        payment_methods = []
        for stat in payment_stats:
            method_name = dict(Booking.PAYMENT_METHOD_CHOICES).get(stat['payment_method'], stat['payment_method'])
            payment_methods.append({
                'name': method_name,
                'value': float(stat['total_amount']),
                'count': stat['count']
            })
        
        return Response({
            'payment_methods': payment_methods,
            'period': period
        })


class BookingStatsView(APIView):
    """Get booking statistics by status"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        period = int(request.GET.get('period', 7))
        shop = request.user.shop
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period)
        
        # Get booking stats by status
        booking_stats = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date
        ).values('booking_status').annotate(
            count=Count('id'),
            revenue=Sum('total_amount')
        ).order_by('-count')
        
        # Format data
        stats = []
        for stat in booking_stats:
            status_name = dict(Booking.BOOKING_STATUS_CHOICES).get(stat['booking_status'], stat['booking_status'])
            stats.append({
                'status': status_name,
                'count': stat['count'],
                'revenue': float(stat['revenue'] or 0)
            })
        
        return Response({
            'booking_stats': stats,
            'period': period
        })


class CustomerAnalyticsView(APIView):
    """Get customer analytics data"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        period = int(request.GET.get('period', 7))
        shop = request.user.shop
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period)
        
        # Get customer stats
        customer_stats = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date
        ).values('user').annotate(
            bookings=Count('id'),
            total_spent=Sum('total_amount', filter=Q(booking_status='completed'))
        ).order_by('-total_spent')
        
        # Get top customers
        top_customers = []
        for stat in customer_stats[:10]:  # Top 10 customers
            user = CustomUser.objects.get(id=stat['user'])
            top_customers.append({
                'name': user.get_full_name() or user.username,
                'email': user.email,
                'bookings': stat['bookings'],
                'total_spent': float(stat['total_spent'] or 0)
            })
        
        # Get new vs returning customers
        total_customers = customer_stats.count()
        new_customers = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date
        ).values('user').annotate(
            first_booking=Min('created_at')
        ).filter(
            first_booking__gte=timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
        ).count()
        
        returning_customers = total_customers - new_customers
        
        return Response({
            'top_customers': top_customers,
            'total_customers': total_customers,
            'new_customers': new_customers,
            'returning_customers': returning_customers,
            'period': period
        })


class HourlyBookingStatsView(APIView):
    """Get hourly booking statistics"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        period = int(request.GET.get('period', 7))
        shop = request.user.shop
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period)
        
        # Get hourly booking data
        bookings = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date,
            booking_status__in=['completed', 'confirmed']
        ).extra(
            select={'hour': "EXTRACT(hour FROM appointment_time)"}
        ).values('hour').annotate(
            count=Count('id'),
            revenue=Sum('total_amount')
        ).order_by('hour')
        
        # Format data for chart
        hourly_data = []
        for booking in bookings:
            hour = int(booking['hour'])
            time_label = f"{hour:02d}:00"
            hourly_data.append({
                'hour': time_label,
                'bookings': booking['count'],
                'revenue': float(booking['revenue'] or 0)
            })
        
        return Response({
            'hourly_data': hourly_data,
            'period': period
        })


class ExportSalesReportView(APIView):
    """Export sales report as PDF or Excel"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        period = int(request.GET.get('period', 7))
        format_type = request.GET.get('format', 'pdf')
        shop = request.user.shop
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=period)
        
        # Get comprehensive data
        bookings = Booking.objects.filter(
            shop=shop,
            appointment_date__gte=start_date,
            appointment_date__lte=end_date
        ).select_related('user').prefetch_related('services')
        
        if format_type == 'pdf':
            return self._generate_pdf_report(bookings, shop, start_date, end_date, period)
        elif format_type == 'excel':
            return self._generate_excel_report(bookings, shop, start_date, end_date, period)
        
        return Response({'error': 'Invalid format'}, status=status.HTTP_400_BAD_REQUEST)
    
    def _generate_pdf_report(self, bookings, shop, start_date, end_date, period):
        """Generate PDF report"""
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph(f"Sales Report - {shop.name}", title_style))
        story.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary stats
        total_revenue = bookings.filter(booking_status='completed').aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        total_bookings = bookings.count()
        completed_bookings = bookings.filter(booking_status='completed').count()
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Revenue', f"₹{total_revenue:,.2f}"],
            ['Total Bookings', str(total_bookings)],
            ['Completed Bookings', str(completed_bookings)],
            ['Completion Rate', f"{(completed_bookings/total_bookings*100):.1f}%" if total_bookings > 0 else "0%"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Bookings table
        story.append(Paragraph("Booking Details", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        booking_data = [['Date', 'Customer', 'Services', 'Amount', 'Status']]
        
        for booking in bookings:
            services_list = ', '.join([service.name for service in booking.services.all()])
            booking_data.append([
                booking.appointment_date.strftime('%Y-%m-%d'),
                booking.user.get_full_name() or booking.user.username,
                services_list[:30] + '...' if len(services_list) > 30 else services_list,
                f"₹{booking.total_amount}",
                booking.get_booking_status_display()
            ])
        
        booking_table = Table(booking_data, colWidths=[1*inch, 1.5*inch, 2*inch, 1*inch, 1*inch])
        booking_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8)
        ]))
        
        story.append(booking_table)
        
        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{period}days.pdf"'
        response.write(pdf_data)
        return response
    
    def _generate_excel_report(self, bookings, shop, start_date, end_date, period):
        """Generate Excel report"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Header
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Sales Report - {shop.name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f"Period: {start_date} to {end_date}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Summary
        ws['A4'] = "Summary"
        ws['A4'].font = Font(size=14, bold=True)
        
        total_revenue = bookings.filter(booking_status='completed').aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        ws['A5'] = "Total Revenue:"
        ws['B5'] = f"₹{total_revenue:,.2f}"
        ws['A6'] = "Total Bookings:"
        ws['B6'] = bookings.count()
        ws['A7'] = "Completed Bookings:"
        ws['B7'] = bookings.filter(booking_status='completed').count()
        
        # Bookings details
        ws['A9'] = "Booking Details"
        ws['A9'].font = Font(size=14, bold=True)
        
        headers = ['Date', 'Customer', 'Services', 'Amount', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, booking in enumerate(bookings, 11):
            ws.cell(row=row, column=1, value=booking.appointment_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=booking.user.get_full_name() or booking.user.username)
            ws.cell(row=row, column=3, value=', '.join([service.name for service in booking.services.all()]))
            ws.cell(row=row, column=4, value=float(booking.total_amount))
            ws.cell(row=row, column=5, value=booking.get_booking_status_display())
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="sales_report_{period}days.xlsx"'
        return response
    




# Add this to your Django views
class ShopSpecificPaymentView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, shop_id):
        """List all commission payments for a specific shop"""
        try:
            # Check if shop exists
            shop = Shop.objects.get(id=shop_id)
            
            # Get query parameters for filtering
            payment_method = request.query_params.get('payment_method')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            # Start with payments for this shop only
            payments = ShopCommissionPayment.objects.filter(shop_id=shop_id)
            
            # Apply additional filters
            if payment_method:
                payments = payments.filter(payment_method=payment_method)
            if start_date:
                payments = payments.filter(payment_date__gte=start_date)
            if end_date:
                payments = payments.filter(payment_date__lte=end_date)
            
            # Order by most recent first
            payments = payments.order_by('-payment_date')
            
            # Calculate total amount
            total_amount = payments.aggregate(
                total=models.Sum('amount')
            )['total'] or 0
            
            # Serialize the data
            payment_data = []
            for payment in payments:
                payment_data.append({
                    'id': payment.id,
                    'shop_id': payment.shop.id,
                    'shop_name': payment.shop.name,
                    'amount': float(payment.amount),
                    'payment_method': payment.payment_method,
                    'transaction_reference': payment.transaction_reference,
                    'payment_date': payment.payment_date.strftime('%Y-%m-%d'),
                    'notes': payment.notes,
                    'paid_by': payment.paid_by.username if payment.paid_by else None,
                    'created_at': payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(payment, 'created_at') else None
                })
            
            return Response({
                'shop': {
                    'id': shop.id,
                    'name': shop.name
                },
                'payments': payment_data,
                'total_count': len(payment_data),
                'total_amount': float(total_amount)
            })
            
        except Shop.DoesNotExist:
            return Response(
                {'error': 'Shop not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

